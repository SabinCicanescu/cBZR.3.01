import os
import pandas as pd
from openpyxl import load_workbook

# Configuration
source_folder = 'path/to/excel/files'  # Folder containing source Excel files
master_file = 'path/to/master.xlsx'    # Path to the master Excel file
sheet_name = 'Sheet1'                  # Sheet name to read/write

# Load existing master file to find the last row
if os.path.exists(master_file):
    book = load_workbook(master_file)
    writer = pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book
    if sheet_name in book.sheetnames:
        last_row = book[sheet_name].max_row
    else:
        last_row = 0
else:
    writer = pd.ExcelWriter(master_file, engine='openpyxl')
    last_row = 0

# Loop through all Excel files in the folder
for filename in os.listdir(source_folder):
    if filename.endswith('.xlsx') and filename != os.path.basename(master_file):
        file_path = os.path.join(source_folder, filename)
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Append data starting from the last row
        df.to_excel(writer, sheet_name=sheet_name, startrow=last_row, index=False, header=False)
        last_row += len(df)

# Save changes
writer.close()
print("Data merged successfully into master file.")

